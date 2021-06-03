# coding = utf-8
import openpyxl
import matplotlib.pyplot as plt

area_man_womanlist = []  ##[[地区, 男性人数, 女性人数], []...... []]
manrow = 6   #男性所在列
womanrow = 7 #女性所在列
startrow = 9 #有效数字开始的行


def loadinfo(fname):
    try:
        openpyxl.load_workbook(fname) 
    except FileNotFoundError:
        print(">>>>>> 没有该文件 <<<<<<")
        return
    infotable = openpyxl.load_workbook(fname)
    allsheets = infotable.sheetnames   #所有工作表
    nowsheet = infotable.worksheets[0] #当前工作表 index从0开始
    # title = nowsheet.title # 工作表名称
    ##行和列
    nrows = nowsheet.max_row
    ncols = nowsheet.max_column
    for i in range(startrow,nrows+1):
        if nowsheet.cell(i,1).value != None:
            area_man_womanlist.append([ str(nowsheet.cell(i,1).value).replace(" ", ""), nowsheet.cell(i,manrow).value, nowsheet.cell(i,womanrow).value])
        else:
            return

##导入数据
loadinfo("A0101a.xlsx")     

plt.rcParams['font.sans-serif'] = ['SimHei']  #用来正常显示中文标签
womans = [woman[2] for woman in area_man_womanlist]
mans = [man[1] for man in area_man_womanlist]
areas = [area[0] for area in area_man_womanlist]

# 根据数据绘制图形
fig = plt.figure(dpi=80, figsize=(10, 6))
width = 0.45
mx = [i for i in range(0,len(area_man_womanlist))]
plt.bar(mx, mans, width = width, label='男性')
lx = [i + width/2 for i in mx]
plt.xticks(lx,areas)
wx = [i + width for i in mx]
plt.bar(wx, womans,width = width, label='女性')
plt.legend() ##添加图例

# 设置图形的格式
plt.xticks(rotation=0) ##角度
plt.title("全国各地区男女人口数", fontsize=20)
plt.xlabel("地区", fontsize=15)
plt.ylabel("人口数", fontsize=15)

# 设置数字标签
for a, b in zip(mx, mans):
    plt.text(a, b, round((b/10**7), 2), ha='center', va='bottom', fontsize=8)
for a, b in zip(wx, womans):
    plt.text(a, b, round((b/10**7), 2), ha='center', va='bottom', fontsize=8)

plt.show()